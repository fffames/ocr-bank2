"""Excel export service for OCR Bank receipts."""
from datetime import datetime
from typing import List, Dict, Any
from sqlalchemy.orm import Session

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, NumberFormat
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from app.models.receipt import Receipt


class ExcelExportService:
    """Service for exporting receipts to Excel files."""

    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl library not installed. Run: pip install openpyxl")

    def export_receipts(
        self,
        receipts: List[Receipt],
        user_timezone: str = "Asia/Bangkok"
    ) -> bytes:
        """
        Export receipts to Excel file format.

        Args:
            receipts: List of receipts to export
            user_timezone: User's timezone for date formatting

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        # Separate receipts by transaction type
        payments = [r for r in receipts if r.transaction_type == "sending"]
        income = [r for r in receipts if r.transaction_type == "receiving"]

        # Create Payments sheet
        if payments:
            self._create_sheet(wb, payments, "Payments")

        # Create Income sheet
        if income:
            self._create_sheet(wb, income, "Income")

        # Remove default Summary sheet if empty
        if not payments and not income:
            wb.remove(ws)

        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def _create_sheet(self, wb: Workbook, receipts: List[Receipt], sheet_name: str):
        """
        Create a sheet with receipt data.

        Args:
            wb: Excel workbook
            receipts: List of receipts for this sheet
            sheet_name: Name of the sheet
        """
        ws = wb.create_sheet(title=sheet_name)

        # Define headers
        headers = ["Date", "Sender/Receiver", "Amount (THB)", "Note"]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_num, receipt in enumerate(receipts, 2):
            # Date
            if receipt.extracted_date:
                ws.cell(row=row_num, column=1, value=receipt.extracted_date.isoformat())

            # Sender/Receiver
            if receipt.transaction_type == "sending":
                ws.cell(row=row_num, column=2, value=receipt.receiver or "")
            else:  # receiving
                ws.cell(row=row_num, column=2, value=receipt.sender or "")

            # Amount
            if receipt.amount:
                cell = ws.cell(row=row_num, column=3, value=float(receipt.amount))
                cell.number_format = '#,##0.00'

            # Note
            ws.cell(row=row_num, column=4, value=receipt.note or "")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].letter  # Get column letter (A, B, C, etc.)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


# Singleton instance
_excel_export_service = None


def get_excel_export_service() -> ExcelExportService:
    """Get or create Excel export service singleton."""
    global _excel_export_service
    if _excel_export_service is None:
        _excel_export_service = ExcelExportService()
    return _excel_export_service
